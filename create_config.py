import openpyxl
from openpyxl import Workbook

def create_config_xlsx():
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Config"

    # Define the configuration categories and their values
    config = {
        "Checkpoints": {
            "ckpt_name": "meinamix_v12Final.safetensors"
        },
        "LoRA_1": {
            "name": "Blue_Future.safetensors",
            "strength_model": 1,
            "strength_clip": 1
        },
        "LoRA_2": {
            "name": "WaterColoursScenev1.0.safetensors",
            "strength_model": 0,
            "strength_clip": 0
        },
        "KSampler_1": {
            "cfg": 7,
            "steps": 20
        },
        "KSampler_2": {
            "cfg": 7
        },
        "Resolution": {
            "width": 912,
            "height": 512
        },
        "Upscale_Resolution": {
            "width": 3840,
            "height": 2160
        }
    }

    # Write headers
    ws['A1'] = 'Category'
    ws['B1'] = 'Parameter'
    ws['C1'] = 'Value'

    # Write configuration data
    row = 2
    for category, params in config.items():
        for param, value in params.items():
            ws[f'A{row}'] = category
            ws[f'B{row}'] = param
            ws[f'C{row}'] = value
            row += 1

    # Save the workbook
    wb.save('config.xlsx')
    print("Configuration file 'config.xlsx' has been created.")

if __name__ == "__main__":
    create_config_xlsx() 